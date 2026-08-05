"""Adiciona coluna 'Imagens site (assets)' na planilha master (Drive .xlsx).

Fluxo: baixa master fresca -> adiciona coluna por aba de fornecedor ->
upload PATCH -> reparo do cache de fórmulas (copy->Sheets->export->PATCH->delete).
Fonte dos dados: audit_imagens.json (auditoria site x assets desta sessão).
"""
import sys, os, json, urllib.request, urllib.parse

sys.path.insert(0, '/Users/guilhermeferreira/Documents/DEV/shopify-agro-pecas-theme/util/master-produtos/integra')
import pipeline as P
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SHEET_ID = P.SHEET_ID
SUP = P.SUP + ['08_FERTISYSTEM']
COL_HDR = 'Imagens site (assets)'

audit = json.load(open(os.path.join(HERE, "..", "relatorios", "audit_imagens_site_20260804.json")))
site = {}
for r in audit:
    for k in {r['sku'], r['sku'][:-2] if r['sku'].endswith('.0') else r['sku'], r['sku'] + '.0'}:
        if k: site[k] = r

def req(url, data=None, method='GET', headers=None, raw=False):
    h = {'Authorization': 'Bearer ' + P.gtoken()}
    h.update(headers or {})
    rq = urllib.request.Request(url, data=data, headers=h, method=method)
    with urllib.request.urlopen(rq, timeout=120) as resp:
        b = resp.read()
    return b if raw else (json.loads(b) if b else {})

def lookup(sku, kit):
    sku = str(sku).strip()
    if kit:
        try:
            k = int(float(kit))
            if k > 1 and f'{sku}-KIT{k}' in site:
                return site[f'{sku}-KIT{k}']
        except (ValueError, TypeError):
            pass
    return site.get(sku)

def main():
    src = os.path.join(HERE, 'master_fresh.xlsx')
    b = req(f'https://www.googleapis.com/drive/v3/files/{SHEET_ID}?alt=media&supportsAllDrives=true', raw=True)
    open(src, 'wb').write(b)
    print('baixada', len(b), 'bytes', flush=True)

    wb = openpyxl.load_workbook(src)  # data_only=False: preserva fórmulas
    preenchidas = 0
    for name in SUP:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        hr = 5
        hdrs = {c: str(ws.cell(hr, c).value) for c in range(1, ws.max_column + 1) if ws.cell(hr, c).value}
        col_img = next((c for c, h in hdrs.items() if h == COL_HDR), max(hdrs) + 1)
        ws.cell(hr, col_img, COL_HDR)
        col_kit = next((c for c, h in hdrs.items() if 'kit' in h.lower() and 'peso' not in h.lower()), None)
        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            sku = ws.cell(r, 1).value
            if sku in (None, ''): continue
            info = lookup(sku, ws.cell(r, col_kit).value if col_kit else None)
            if info is None:
                val = ''
            elif info['imagens']:
                val = ' | '.join(info['imagens'])
            elif 'Bomba' in (info.get('title') or ''):
                # fallback definido 04/08: bombas sem foto usam a imagem genérica do tema
                val = 'bomba_generica.png (genérica)'
            else:
                val = 'SEM IMAGEM'
            if val:
                ws.cell(r, col_img, val); n += 1
        preenchidas += n
        print(f'{name}: col {col_img}, {n} linhas preenchidas', flush=True)
    print('total preenchidas:', preenchidas, flush=True)

    out = os.path.join(HERE, 'master_updated.xlsx')
    wb.save(out)

    MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    req(f'https://www.googleapis.com/upload/drive/v3/files/{SHEET_ID}?uploadType=media&supportsAllDrives=true',
        data=open(out, 'rb').read(), method='PATCH', headers={'Content-Type': MIME})
    print('upload OK', flush=True)

    # reparo do cache de fórmulas: copy->Google Sheets (recalcula)->export xlsx->PATCH->delete
    cp = req(f'https://www.googleapis.com/drive/v3/files/{SHEET_ID}/copy?supportsAllDrives=true',
             data=json.dumps({'mimeType': 'application/vnd.google-apps.spreadsheet',
                              'name': 'tmp_recalc_master'}).encode(),
             method='POST', headers={'Content-Type': 'application/json'})
    cid = cp['id']
    print('cópia recalculada', cid, flush=True)
    xb = req(f'https://www.googleapis.com/drive/v3/files/{cid}/export?mimeType={urllib.parse.quote(MIME)}', raw=True)
    req(f'https://www.googleapis.com/upload/drive/v3/files/{SHEET_ID}?uploadType=media&supportsAllDrives=true',
        data=xb, method='PATCH', headers={'Content-Type': MIME})
    req(f'https://www.googleapis.com/drive/v3/files/{cid}?supportsAllDrives=true', method='DELETE')
    print('cache de fórmulas reparado e cópia removida — planilha master atualizada.', flush=True)

if __name__ == '__main__':
    main()
