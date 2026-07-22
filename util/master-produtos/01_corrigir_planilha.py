#!/usr/bin/env python3
"""
01 — Saneamento da planilha master + unificação de fórmulas.

O que faz (em todas as abas de produto):
  1. Converte números-texto ('R$ 19,10', '0,36', '87,5') em números reais.
  2. Cria a célula global 01_CONFIG_GLOBAL!C14 = 0,2082 (taxa MP 5% + parcelado 16,82%)
     e aponta a coluna Taxa MP de TODAS as abas para ela.
  3. Reescreve a cadeia de fórmulas de cada linha (cubado, considerado, custo+frete,
     preço c/ margem, taxa, preço final) com referências corretas da própria linha.
  4. Linhas kit (Qtde KIT > 1): status/margem herdados da linha unitária, custo e peso
     viram fórmulas referenciando a linha unitária (=custo_unit × N); caixas com dims
     iguais à peça unitária são re-estimadas a partir de outra caixa de kit do mesmo SKU.
  5. Adiciona colunas derivadas 'SKU Shopify (derivado)' e 'Título Shopify (derivado)'.
  6. Flags de pendência em Observações (STARA sem peso, linhas ADG incompletas,
     duplicado 6237989M1) e harmonização de margens (0,19999→0,2; 6237989M1 → 0,2).

Uso:  python3 01_corrigir_planilha.py [--dry]
"""
import csv, sys, datetime
import openpyxl
from openpyxl.utils import get_column_letter as L
from common import (XLSX, REL_DIR, PRODUCT_SHEETS, HEADER_ROW, DATA_START,
                    TAXA_MP, colmap, iter_data_rows, parse_br_number)

DRY = "--dry" in sys.argv
TS  = datetime.datetime.now().strftime("%Y%m%d_%H%M")
LOG = []

def log(sheet, row, field, old, new, motivo):
    LOG.append({"aba": sheet, "linha": row, "campo": field,
                "antes": repr(old), "depois": repr(new), "motivo": motivo})

NUMERIC_KEYS = ["kit", "lead", "peso", "comp", "larg", "alt", "frete", "custo", "marg"]
INT_KEYS = {"kit", "lead"}

def kitqty(ws, cm, r):
    if "kit" not in cm:
        return None
    v = ws.cell(row=r, column=cm["kit"]).value
    n = parse_br_number(v)
    return int(n) if n is not None else None

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=False)

    # ── 2. célula global da taxa MP ─────────────────────────────────────────
    cfg = wb["01_CONFIG_GLOBAL"]
    old = cfg["C14"].value
    cfg["B14"] = "Taxa MP efetiva (5% cartão + parcelado 12x 16,82%)"
    cfg["C14"] = TAXA_MP
    cfg["D14"] = "FONTE ÚNICA da taxa usada no Preço final de TODAS as abas. Decisão 21/07/2026."
    log("01_CONFIG_GLOBAL", 14, "C14", old, TAXA_MP, "taxa MP unificada")
    TAXA_REF = "='01_CONFIG_GLOBAL'!$C$14"

    for name in PRODUCT_SHEETS:
        ws = wb[name]
        cm = colmap(ws)
        rows = list(iter_data_rows(ws, cm))
        if not rows:
            continue

        # ── 1. sanitização texto→número ────────────────────────────────────
        for r, sku in rows:
            for key in NUMERIC_KEYS:
                if key not in cm:
                    continue
                cell = ws.cell(row=r, column=cm[key])
                v = cell.value
                if isinstance(v, str) and not v.startswith("="):
                    n = parse_br_number(v)
                    if n is not None:
                        n = int(n) if key in INT_KEYS and float(n).is_integer() else n
                        cell.value = n
                        log(name, r, key, v, n, "texto→número")
            # margens 0.19999… → 0.2
            mc = ws.cell(row=r, column=cm["marg"])
            if isinstance(mc.value, float) and abs(mc.value - round(mc.value, 2)) > 1e-12:
                log(name, r, "marg", mc.value, round(mc.value, 2), "arredondamento")
                mc.value = round(mc.value, 2)

        # ── índice de linhas unitárias e de kit por SKU ─────────────────────
        unit_row, kit_rows = {}, {}
        for r, sku in rows:
            q = kitqty(ws, cm, r)
            if q is None or q <= 1:
                unit_row.setdefault(sku, r)
            else:
                kit_rows.setdefault(sku, []).append(r)

        # ── 4. normalização das linhas kit ──────────────────────────────────
        for sku, krs in kit_rows.items():
            u = unit_row.get(sku)
            for r in krs:
                q = kitqty(ws, cm, r)
                kcol = L(cm["kit"])
                # status herdado
                st = ws.cell(row=r, column=cm["status"]).value
                if st in (None, "") and u:
                    ust = ws.cell(row=u, column=cm["status"]).value
                    if ust:
                        ws.cell(row=r, column=cm["status"]).value = ust
                        log(name, r, "status", st, ust, "herdado da linha unitária")
                # margem herdada
                mg = ws.cell(row=r, column=cm["marg"])
                if mg.value in (None, "") and u:
                    mg.value = f"={L(cm['marg'])}{u}"
                    log(name, r, "marg", None, mg.value, "herdada da linha unitária")
                # custo do kit
                cc = ws.cell(row=r, column=cm["custo"])
                ucost = parse_br_number(ws.cell(row=u, column=cm["custo"]).value) if u else None
                cur = parse_br_number(cc.value) if not (isinstance(cc.value, str) and cc.value.startswith("=")) else "FORMULA"
                if u:
                    want = f"={L(cm['custo'])}{u}*{kcol}{r}"
                    if cc.value in (None, "") or cur == ucost or (isinstance(cc.value, str) and cc.value.startswith("=")):
                        if cc.value != want:
                            log(name, r, "custo", cc.value, want, "custo kit = unit × N")
                            cc.value = want
                elif cc.value in (None, ""):
                    ref = next((k for k in krs if k != r and ws.cell(row=k, column=cm["custo"]).value not in (None, "")), None)
                    if ref:
                        want = f"={L(cm['custo'])}{ref}/{kcol}{ref}*{kcol}{r}"
                        cc.value = want
                        log(name, r, "custo", None, want, "proporcional a outro lote (sem linha unitária)")
                # peso do kit
                pc = ws.cell(row=r, column=cm["peso"])
                upeso = parse_br_number(ws.cell(row=u, column=cm["peso"]).value) if u else None
                pval = parse_br_number(pc.value) if not (isinstance(pc.value, str) and pc.value.startswith("=")) else None
                if u and (pc.value in (None, "") or (isinstance(pc.value, str) and pc.value.startswith("="))
                          or (pval is not None and upeso is not None and abs(pval - upeso) < 1e-9)):
                    want = f"={L(cm['peso'])}{u}*{kcol}{r}"
                    if pc.value != want:
                        log(name, r, "peso", pc.value, want, "peso kit = unit × N")
                        pc.value = want
                # caixa do kit com dims da peça unitária → re-estimar
                if u:
                    dims  = [parse_br_number(ws.cell(row=r, column=cm[k]).value) for k in ("comp", "larg", "alt")]
                    udims = [parse_br_number(ws.cell(row=u, column=cm[k]).value) for k in ("comp", "larg", "alt")]
                    if None not in dims and None not in udims and all(abs(a-b) < 1e-9 for a, b in zip(dims, udims)):
                        ref = next((k for k in krs if k != r and
                                    None not in [parse_br_number(ws.cell(row=k, column=cm[x]).value) for x in ("comp","larg","alt")] and
                                    [parse_br_number(ws.cell(row=k, column=cm[x]).value) for x in ("comp","larg","alt")] != udims), None)
                        if ref:
                            qref = kitqty(ws, cm, ref)
                            rd = [parse_br_number(ws.cell(row=ref, column=cm[x]).value) for x in ("comp","larg","alt")]
                            new = [rd[0], rd[1], max(2.0, round(rd[2] * q / qref, 1))]
                            for k, v in zip(("comp", "larg", "alt"), new):
                                ws.cell(row=r, column=cm[k]).value = v
                            ws.cell(row=r, column=cm["origem"]).value = "estimado"
                            log(name, r, "caixa", f"{dims}", f"{new}", f"dims eram da peça unitária — escalado da caixa kit {qref}")

        # ── 3. cadeia de fórmulas (todas as linhas com SKU) ─────────────────
        for r, sku in rows:
            C, Lg, A = L(cm["comp"]), L(cm["larg"]), L(cm["alt"])
            P = L(cm["peso"])
            cub, cons = L(cm["cubado"]), L(cm["considerado"])
            fr, cu, mg = L(cm["frete"]), L(cm["custo"]), L(cm["marg"])
            cf, pcm = L(cm["cf"]), L(cm["pcm"])
            tx, txr, fi = L(cm["taxa"]), L(cm["taxa_rs"]), L(cm["final"])

            def setf(col, formula, field):
                cell = ws.cell(row=r, column=cm[col])
                if cell.value != formula:
                    cell.value = formula
            if "cubado_kit" in cm:
                setf("cubado_kit", f"=IFERROR(IF(AND({C}{r}>0,{Lg}{r}>0,{A}{r}>0),({C}{r}*{Lg}{r}*{A}{r})/6000,\"\"),\"\")", "cubado_kit")
            setf("cubado", f"=IFERROR(IF(AND({C}{r}>0,{Lg}{r}>0,{A}{r}>0),({C}{r}*{Lg}{r}*{A}{r})/6000,\"\"),\"\")", "cubado")
            setf("considerado", f"=IFERROR(IF(AND({P}{r}>0,{cub}{r}>0),MAX({P}{r},{cub}{r}),IF({P}{r}>0,{P}{r},IF({cub}{r}>0,{cub}{r},\"\"))),\"\")", "considerado")
            setf("cf", f"=IFERROR({cu}{r}+{fr}{r},\"\")", "cf")
            setf("pcm", f"=IFERROR(IF({mg}{r}<1,{cf}{r}/(1-{mg}{r}),\"\"),\"\")", "pcm")
            tcell = ws.cell(row=r, column=cm["taxa"])
            if tcell.value != TAXA_REF:
                log(name, r, "taxa", tcell.value, TAXA_REF, "ref. única CONFIG!C14")
                tcell.value = TAXA_REF
            setf("taxa_rs", f"=IFERROR({pcm}{r}*{tx}{r},\"\")", "taxa_rs")
            setf("final", f"=IFERROR(IF({tx}{r}<1,{pcm}{r}/(1-{tx}{r}),\"\"),\"\")", "final")

        # cabeçalho da taxa unificado
        ws.cell(row=HEADER_ROW, column=cm["taxa"]).value = "Taxa Mercado Pago (%)"

        # ── 5. colunas derivadas ────────────────────────────────────────────
        last_hdr = max(c for c in range(1, ws.max_column + 1)
                       if ws.cell(row=HEADER_ROW, column=c).value is not None)
        cm2 = colmap(ws)
        if "sku_deriv" not in cm2:
            cS, cT = last_hdr + 1, last_hdr + 2
            ws.cell(row=HEADER_ROW, column=cS).value = "SKU Shopify (derivado)"
            ws.cell(row=HEADER_ROW, column=cT).value = "Título Shopify (derivado)"
        else:
            cS, cT = cm2["sku_deriv"], cm2["tit_deriv"]
        for r, sku in rows:
            if "kit" in cm:
                K = L(cm["kit"])
                ws.cell(row=r, column=cS).value = f"=IF(AND({K}{r}<>\"\",{K}{r}>1),A{r}&\"-KIT\"&{K}{r},A{r})"
                ws.cell(row=r, column=cT).value = f"=IF(AND({K}{r}<>\"\",{K}{r}>1),D{r}&\" — Kit \"&{K}{r}&\" unidades\",D{r})"
            else:
                ws.cell(row=r, column=cS).value = f"=A{r}"
                ws.cell(row=r, column=cT).value = f"=D{r}"

        # ── 6. pendências em Observações ────────────────────────────────────
        def obs_append(r, txt):
            cell = ws.cell(row=r, column=cm["obs"])
            cur = str(cell.value or "").strip()
            if txt not in cur:
                cell.value = (cur + " | " if cur else "") + txt
                log(name, r, "obs", cur, cell.value, "pendência")

        for r, sku in rows:
            st = str(ws.cell(row=r, column=cm["status"]).value or "").lower()
            title = ws.cell(row=r, column=cm["title"]).value
            custo = ws.cell(row=r, column=cm["custo"]).value
            peso  = ws.cell(row=r, column=cm["peso"]).value
            if st == "active" and (title in (None, "") or custo in (None, "")):
                obs_append(r, "PENDÊNCIA: linha incompleta (sem título/custo) — fora do sync")
            elif st == "active" and peso in (None, "") and kitqty(ws, cm, r) in (None, 0, 1):
                obs_append(r, "PENDÊNCIA: sem peso/dimensões — frete default R$25 (estimado); pesar/medir p/ recotar")
            if sku == "6237989M1" and name == "06_STARA":
                mc = ws.cell(row=r, column=cm["marg"])
                if parse_br_number(mc.value) not in (None,) and not str(mc.value).startswith("="):
                    if abs((parse_br_number(mc.value) or 0) - 0.2) > 1e-9:
                        log(name, r, "marg", mc.value, 0.2, "harmonizado com 03_AGCO")
                        mc.value = 0.2
                obs_append(r, "DUPLICADO com 03_AGCO (canônico lá) — esta linha não sobe no sync")

    wb.calculation.fullCalcOnLoad = True
    if DRY:
        print(f"[DRY] {len(LOG)} alterações NÃO gravadas")
    else:
        wb.save(XLSX)
        print(f"Planilha gravada. {len(LOG)} alterações registradas.")
    REL_DIR.mkdir(exist_ok=True)
    out = REL_DIR / f"correcoes_{TS}{'_dry' if DRY else ''}.csv"
    with open(out, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["aba", "linha", "campo", "antes", "depois", "motivo"])
        w.writeheader(); w.writerows(LOG)
    print(f"Log: {out}")
    # resumo por motivo
    from collections import Counter
    for motivo, n in Counter(x["motivo"] for x in LOG).most_common():
        print(f"  {n:4d} × {motivo}")

if __name__ == "__main__":
    main()
