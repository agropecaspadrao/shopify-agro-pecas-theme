#!/usr/bin/env python3
"""
Base compartilhada do pipeline master-produtos.
Lê o .env da raiz, mapeia colunas por cabeçalho (linha 5) e expõe helpers
de normalização de SKU e cálculo da cadeia de preço.
"""
import json, pathlib, re, unicodedata
from urllib.request import urlopen, Request
from urllib.parse import urlencode

BASE_DIR   = pathlib.Path(__file__).parent
REPO_DIR   = BASE_DIR.parent.parent
ENV_FILE   = REPO_DIR / ".env"
XLSX       = REPO_DIR / "util" / "APP_Master_Produtos_Shopify.xlsx"
REL_DIR    = BASE_DIR / "relatorios"

API_VERSION = "2025-01"
TINY_BASE   = "https://api.tiny.com.br/api2"
FRENET_URL  = "https://api.frenet.com.br/shipping/quote"

CEP_DESTINO = "81220310"          # estoque APP — Rua José Benedito Cottolengo 901, Curitiba/PR
TAXA_MP     = 0.2082              # 5% cartão + parcelado 12x 16,82% — decisão 21/07/2026

# Origem do frete inbound por aba (decisão 21/07/2026): Greco cota da sede Greco;
# AGCO/JD/STARA/GTS/FERTI cotam da ADG Plásticos. SOHIPREN não cota (frete já preenchido).
CEP_ORIGEM = {
    "04_GRECO":      "99025000",  # Agro e Greco Agro Tech — Av. Brasil Oeste 560, Passo Fundo/RS
    "03_AGCO":       "99051380",  # ADG Plásticos — Rua Gaspar Martins 2035, Passo Fundo/RS
    "05_JOHN_DEERE": "99051380",
    "06_STARA":      "99051380",
    "07_GTS":        "99051380",
    "08_FERTISYSTEM":"99051380",
}
FRETE_DEFAULT = 25.0              # CONFIG C6 — usado quando não há peso/dimensões

# Abas de produto (ordem de processamento) — 08_FERTISYSTEM vazia, mantida por completude
PRODUCT_SHEETS = ["02_SOHIPREN", "04_GRECO", "03_AGCO", "05_JOHN_DEERE",
                  "06_STARA", "07_GTS", "08_FERTISYSTEM"]
NO_QUOTE_SHEETS = {"02_SOHIPREN"}        # frete inbound já preenchido — não recotar

HEADER_ROW = 5
DATA_START = 6

# chave lógica -> fragmento do cabeçalho (case-insensitive)
COL_KEYS = {
    "sku":        "sku / código",
    "part":       "part number",
    "handle":     "handle",
    "title":      "title",
    "type":       "product type",
    "vendor":     "vendor",
    "tags":       "tags",
    "status":     "status",
    "montadora":  "montadora",
    "equip":      "equipamento",
    "funcao":     "função / aplicação",
    "cruzados":   "códigos cruzados",
    "ncm":        "ncm",
    "kit":        "qtde kit",
    "lead":       "lead time",
    "peso":       "peso (kg)",          # exato — evita colidir com cubado/considerado
    "comp":       "comprimento",
    "larg":       "largura",
    "alt":        "altura",
    "cubado_kit": "peso cubado (kg) kit",
    "cubado":     "peso cubado (kg)",
    "considerado":"peso considerado",
    "frete":      "frete fábrica",
    "origem":     "origem do dado",
    "custo":      "custo unitário",
    "marg":       "% margem",
    "taxa":       "taxa mercado pago",
    "taxa_rs":    "taxa mp ()",
    "cf":         "custo + frete",
    "pcm":        "preço c/ margem",
    "final":      "preço final",
    "cad":        "cadastrado shopify",
    "obs":        "observações",
    "sku_deriv":  "sku shopify (derivado)",
    "tit_deriv":  "título shopify (derivado)",
}

def load_env():
    env = {}
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.split("#")[0].strip().strip('"').strip("'") if k.strip() != "CLIENT_SCREDT_GCP_ADS" else v
    return env

FRENET_TOKEN = load_env().get("FRENET_TOKEN", "")

def colmap(ws):
    """chave lógica -> índice de coluna (1-based), por matching de cabeçalho da linha 5."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v is not None:
            headers[c] = str(v).strip().lower()
    out = {}
    # 'peso (kg)' precisa ser exato p/ não pegar 'peso cubado (kg)'
    for key, frag in COL_KEYS.items():
        for c, h in headers.items():
            if key == "peso":
                if h == "peso (kg)":
                    out[key] = c; break
            elif key == "cubado":
                if h == "peso cubado (kg)":
                    out[key] = c; break
            elif frag in h:
                out[key] = c; break
    return out

def iter_data_rows(ws, cm):
    for r in range(DATA_START, ws.max_row + 1):
        sku = ws.cell(row=r, column=cm["sku"]).value
        if sku not in (None, ""):
            yield r, str(sku).strip()

BR_NUM = re.compile(r"^\s*(?:R\$)?\s*(-?[\d.,]+)\s*$")

def parse_br_number(v):
    """'R$ 19,10' -> 19.10 · '0,36' -> 0.36 · '1.234,56' -> 1234.56 · retorna None se não numérico."""
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return None
    m = BR_NUM.match(v)
    if not m:
        return None
    s = m.group(1)
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def norm_sku(s):
    """Normalização p/ matching entre planilha, Shopify e Tiny."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().upper().strip()
    return re.sub(r"[^A-Z0-9]", "", s)

def sku_match_keys(s):
    """Chaves de match em ordem de prioridade (exata normalizada, sem sufixo .0)."""
    s = str(s).strip()
    keys = [norm_sku(s)]
    if s.endswith(".0"):
        keys.append(norm_sku(s[:-2]))
    return keys

def preco_final(custo, frete, margem, taxa=TAXA_MP):
    """(custo+frete)/(1-margem)/(1-taxa) — espelho da cadeia da planilha."""
    if custo is None or margem is None or margem >= 1 or taxa >= 1:
        return None
    cf = custo + (frete or 0.0)
    return round(cf / (1 - margem) / (1 - taxa), 2)

# ── Leitura resolvida da planilha ────────────────────────────────────────────
_REF = re.compile(r"^=([A-Z]{1,2})(\d+)$")                      # =Y6
_MUL = re.compile(r"^=([A-Z]{1,2})(\d+)\*([A-Z]{1,2})(\d+)$")   # =Y6*N10
_PROP = re.compile(r"^=([A-Z]{1,2})(\d+)/([A-Z]{1,2})(\d+)\*([A-Z]{1,2})(\d+)$")  # =X12/M12*M37

def _resolve(ws, col_letter, row, depth=0):
    """Resolve célula numérica ou fórmula simples (ref, a*b, a/b*c) — espelho do Excel."""
    from openpyxl.utils import column_index_from_string
    if depth > 4:
        return None
    v = ws.cell(row=row, column=column_index_from_string(col_letter)).value
    n = parse_br_number(v)
    if n is not None:
        return float(n)
    if not isinstance(v, str):
        return None
    m = _REF.match(v)
    if m:
        return _resolve(ws, m.group(1), int(m.group(2)), depth + 1)
    m = _MUL.match(v)
    if m:
        a = _resolve(ws, m.group(1), int(m.group(2)), depth + 1)
        b = _resolve(ws, m.group(3), int(m.group(4)), depth + 1)
        return a * b if a is not None and b is not None else None
    m = _PROP.match(v)
    if m:
        a = _resolve(ws, m.group(1), int(m.group(2)), depth + 1)
        b = _resolve(ws, m.group(3), int(m.group(4)), depth + 1)
        c = _resolve(ws, m.group(5), int(m.group(6)), depth + 1)
        return a / b * c if None not in (a, b, c) and b else None
    # fórmula literal antiga tipo =0.074*M7
    m = re.match(r"^=([\d.]+)\*([A-Z]{1,2})(\d+)$", v)
    if m:
        b = _resolve(ws, m.group(2), int(m.group(3)), depth + 1)
        return float(m.group(1)) * b if b is not None else None
    return None

def load_master(wb):
    """Lê todas as linhas com SKU das abas de produto, com valores resolvidos.
    Retorna lista de dicts. Linhas 'active' incompletas (sem título/custo) são marcadas skip."""
    from openpyxl.utils import get_column_letter as L
    out = []
    for name in PRODUCT_SHEETS:
        ws = wb[name]
        cm = colmap(ws)
        if "sku" not in cm:
            continue
        rows = list(iter_data_rows(ws, cm))
        for r, sku in rows:
            g = lambda k: ws.cell(row=r, column=cm[k]).value if k in cm else None
            num = lambda k: _resolve(ws, L(cm[k]), r) if k in cm else None
            kit = num("kit")
            kit = int(kit) if kit else None
            d = {
                "sheet": name, "row": r, "sku": sku,
                "part": g("part"), "title": g("title"), "type": g("type"),
                "vendor": g("vendor"), "tags": g("tags"),
                "status": str(g("status") or "").strip().lower(),
                "montadora": g("montadora"), "equip": g("equip"),
                "funcao": g("funcao") if not str(g("funcao") or "").startswith("=") else None,
                "cruzados": g("cruzados"), "ncm": g("ncm"), "lead": num("lead"),
                "kit": kit, "is_kit": bool(kit and kit > 1),
                "peso": num("peso"), "comp": num("comp"), "larg": num("larg"), "alt": num("alt"),
                "frete": num("frete"), "origem": g("origem"),
                "custo": num("custo"), "marg": num("marg"),
                "cad": g("cad"), "obs": str(g("obs") or ""),
            }
            d["sku_shopify"] = f"{sku}-KIT{kit}" if d["is_kit"] else sku
            base_title = str(d["title"] or "").strip()
            d["titulo_shopify"] = (f"{base_title} — Kit {kit} unidades" if d["is_kit"] else base_title)
            d["skip"] = d["status"] != "active" or not base_title or d["custo"] is None \
                        or "PENDÊNCIA: linha incompleta" in d["obs"] \
                        or "DUPLICADO com 03_AGCO" in d["obs"]
            out.append(d)
    # dedup por sku_shopify (ex.: KK31823-KIT25 em duas linhas): mantém a primeira ativa
    seen = {}
    for d in out:
        if d["skip"]:
            continue
        k = d["sku_shopify"]
        if k in seen:
            d["skip"] = True
            d["obs"] += " | DUP(sku_shopify) — usa linha " + str(seen[k])
        else:
            seen[k] = d["row"]
    return out

# ── Shopify ──────────────────────────────────────────────────────────────────
def shopify_token(env):
    shop = env["SHOPIFY_SHOP"]
    data = urlencode({"grant_type": "client_credentials",
                      "client_id": env["SHOPIFY_CLIENT_ID"],
                      "client_secret": env["SHOPIFY_CLIENT_SECRET"]}).encode()
    req = Request(f"https://{shop}.myshopify.com/admin/oauth/access_token", data=data, method="POST")
    with urlopen(req, timeout=30) as r:
        return json.loads(r.read())["access_token"]

def shopify_graphql(env, token, query, variables=None):
    shop = env["SHOPIFY_SHOP"]
    body = json.dumps({"query": query, "variables": variables or {}}).encode()
    req = Request(f"https://{shop}.myshopify.com/admin/api/{API_VERSION}/graphql.json",
                  data=body, method="POST",
                  headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"})
    with urlopen(req, timeout=60) as r:
        out = json.loads(r.read())
    if out.get("errors"):
        raise RuntimeError(f"GraphQL errors: {out['errors']}")
    return out["data"]

# ── Tiny ─────────────────────────────────────────────────────────────────────
def tiny_call(env, endpoint, method="GET", **params):
    params["token"] = env["TINY_API_KEY"]
    params["formato"] = "json"
    if method == "GET":
        with urlopen(f"{TINY_BASE}/{endpoint}?" + urlencode(params), timeout=30) as r:
            return json.loads(r.read())
    body = urlencode(params).encode()
    req = Request(f"{TINY_BASE}/{endpoint}", data=body, method="POST")
    with urlopen(req, timeout=30) as r:
        return json.loads(r.read())
