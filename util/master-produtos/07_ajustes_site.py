# -*- coding: utf-8 -*-
"""07 — Ajustes de catálogo no Shopify (decisões de 08/08/2026):

1. Rebrand Greco → APP: título, tags (greco → gps-app) e descrição.
2. Tags de marca do maquinário (agco/stara/gts/gps-app) p/ o filtro lateral.
3. Drafts da planilha master (Stara/GTS/JD/AGCO) → status DRAFT no site.
4. 4 bombas com foto placeholder → imagem única bomba_generica.png.

Uso:
    python3 07_ajustes_site.py           # dry-run (mostra o que faria)
    python3 07_ajustes_site.py --apply   # aplica
"""
import re, sys, time, urllib.request
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from common import load_env, shopify_token, shopify_graphql, colmap, iter_data_rows

BASE = Path(__file__).parent
REPO = BASE.parent.parent
XLSX = REPO / "util" / "APP_Master_Produtos_Shopify.xlsx"
GENERICA = REPO / "assets" / "bomba_generica.png"

APPLY = "--apply" in sys.argv

# bombas que ficam só com a imagem genérica (pedido do Guilherme, 08/08)
GENERIC_SKUS = {"5.1305.0565032.0", "5.0220.0547207", "5.0220.0548817", "5.0209.0547219"}
DRAFT_SHEETS = ["06_STARA", "07_GTS", "05_JOHN_DEERE", "03_AGCO"]

Q = """
query($after: String) {
  products(first: 50, after: $after) {
    pageInfo { hasNextPage endCursor }
    nodes {
      id title status tags descriptionHtml
      variants(first: 1) { nodes { sku } }
      media(first: 15) { nodes { id mediaContentType ... on MediaImage { alt image { url } } } }
    }
  }
}"""

M_UPDATE = """
mutation($input: ProductInput!) {
  productUpdate(input: $input) { product { id } userErrors { field message } }
}"""

M_DEL_MEDIA = """
mutation($productId: ID!, $mediaIds: [ID!]!) {
  productDeleteMedia(productId: $productId, mediaIds: $mediaIds) {
    deletedMediaIds mediaUserErrors { field message }
  }
}"""

M_STAGED = """
mutation($input: [StagedUploadInput!]!) {
  stagedUploadsCreate(input: $input) {
    stagedTargets { url resourceUrl parameters { name value } }
    userErrors { field message }
  }
}"""

M_CREATE_MEDIA = """
mutation($productId: ID!, $media: [CreateMediaInput!]!) {
  productCreateMedia(productId: $productId, media: $media) {
    media { id status } mediaUserErrors { field message }
  }
}"""

M_REORDER = """
mutation($id: ID!, $moves: [MoveInput!]!) {
  productReorderMedia(id: $id, moves: $moves) { job { id } mediaUserErrors { field message } }
}"""

Q_MEDIA_STATUS = """
query($id: ID!) { product(id: $id) { media(first: 20) { nodes { id status } } } }"""


def fetch_all(env, token):
    out, cursor = [], None
    while True:
        d = shopify_graphql(env, token, Q, {"after": cursor})
        p = d["products"]
        out.extend(p["nodes"])
        if not p["pageInfo"]["hasNextPage"]:
            return out
        cursor = p["pageInfo"]["endCursor"]


def brand_tag_for(sku):
    s = sku.upper()
    base = re.sub(r"-KIT\d+$", "", s)
    if base.startswith("IPCX"):
        return "gts"
    if base.startswith(("ACX", "ACW")) or base in {"33070004", "6237989M1"} \
       or base.startswith(("7038105", "7038106", "43839")):
        return "agco"
    if re.match(r"^\d{3,5}-\d{3,4}$", base):
        return "stara"
    if base.startswith("GR"):
        return "gps-app"
    return None


def staged_upload(env, token, filename, data):
    d = shopify_graphql(env, token, M_STAGED, {"input": [{
        "resource": "IMAGE", "filename": filename,
        "mimeType": "image/png", "httpMethod": "POST"}]})
    t = d["stagedUploadsCreate"]
    if t["userErrors"]:
        raise RuntimeError(t["userErrors"])
    tgt = t["stagedTargets"][0]
    boundary = "----appajustes%d" % int(time.time() * 1000)
    parts = []
    for p in tgt["parameters"]:
        parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="{p["name"]}"\r\n\r\n{p["value"]}\r\n'.encode())
    parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="file"; filename="{filename}"\r\nContent-Type: image/png\r\n\r\n'.encode())
    body = b"".join(parts) + data + f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(tgt["url"], data=body, method="POST",
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    with urllib.request.urlopen(req, timeout=120) as r:
        r.read()
    return tgt["resourceUrl"]


def wait_ready(env, token, pid, mid, timeout=60):
    t0 = time.time()
    while time.time() - t0 < timeout:
        d = shopify_graphql(env, token, Q_MEDIA_STATUS, {"id": pid})
        st = {m["id"]: m["status"] for m in d["product"]["media"]["nodes"]}
        if st.get(mid) == "READY":
            return True
        if st.get(mid) == "FAILED":
            return False
        time.sleep(2)
    return False


def planilha_drafts():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=False)
    skus = set()
    for sheet in DRAFT_SHEETS:
        ws = wb[sheet]
        cm = colmap(ws)
        for r, sku in iter_data_rows(ws, cm):
            st = str(ws.cell(row=r, column=cm["status"]).value or "").strip().lower()
            if st != "draft":
                continue
            kit = ws.cell(row=r, column=cm["kit"]).value if "kit" in cm else None
            try:
                kit = int(kit) if kit else 1
            except (TypeError, ValueError):
                kit = 1
            skus.add(sku if kit <= 1 else f"{sku}-KIT{kit}")
    return skus


def main():
    env = load_env()
    token = shopify_token(env)
    prods = fetch_all(env, token)
    by_sku = {}
    for p in prods:
        sku = (p["variants"]["nodes"][0]["sku"] or "").strip() if p["variants"]["nodes"] else ""
        by_sku[sku] = p
    print(f"{len(prods)} produtos")

    def update(pid, inp, label):
        if APPLY:
            d = shopify_graphql(env, token, M_UPDATE, {"input": dict(inp, id=pid)})
            errs = d["productUpdate"]["userErrors"]
            print(("  ✔ " if not errs else f"  ✖ {errs} ") + label)
        else:
            print("  [dry] " + label)

    # ── 1+2. Rebrand Greco→APP e tags de marca ───────────────────────────────
    print("\n— Rebrand Greco → APP + tags de marca —")
    for p in prods:
        sku = (p["variants"]["nodes"][0]["sku"] or "").strip() if p["variants"]["nodes"] else ""
        tags = list(p["tags"])
        newtags = [t for t in tags if t.lower() != "greco"]
        is_greco = "Greco" in p["title"] or len(newtags) != len(tags)
        bt = brand_tag_for(sku)
        if bt and bt not in [t.lower() for t in newtags]:
            newtags.append(bt)
        newtitle = p["title"].replace("Greco Agro Tech", "APP").replace("Greco", "APP")
        inp, changes = {}, []
        if newtitle != p["title"]:
            inp["title"] = newtitle
            changes.append(f"título → {newtitle}")
        if sorted(newtags) != sorted(tags):
            inp["tags"] = newtags
            added = set(newtags) - set(tags)
            removed = set(tags) - set(newtags)
            changes.append(f"tags +{sorted(added)} -{sorted(removed)}")
        if is_greco and p["descriptionHtml"] and "Greco" in p["descriptionHtml"]:
            body = p["descriptionHtml"].replace("Greco Agro Tech", "APP").replace("Greco", "APP")
            inp["descriptionHtml"] = body
            changes.append("descrição sem 'Greco'")
        if inp:
            update(p["id"], inp, f"{sku or p['title'][:30]}: " + "; ".join(changes))

    # ── 3. Drafts da planilha ────────────────────────────────────────────────
    print("\n— Drafts da planilha (Stara/GTS/JD/AGCO) —")
    for sku in sorted(planilha_drafts()):
        p = by_sku.get(sku)
        if p is None:
            continue
        if p["status"] == "ACTIVE":
            update(p["id"], {"status": "DRAFT"}, f"{sku}: ACTIVE → DRAFT ({p['title'][:45]})")

    # ── 4. Bombas com imagem genérica ────────────────────────────────────────
    print("\n— Bombas → bomba_generica.png —")
    for sku in sorted(GENERIC_SKUS):
        p = by_sku.get(sku)
        if p is None:
            print(f"  ! {sku} não encontrado no site")
            continue
        medias = p["media"]["nodes"]
        # remove capas v1 e placeholders; preserva desenho técnico
        to_del = [m["id"] for m in medias
                  if m["mediaContentType"] == "IMAGE"
                  and "desenho_tec" not in (m.get("image") or {}).get("url", "")]
        if not APPLY:
            print(f"  [dry] {sku}: remove {len(to_del)} mídia(s), sobe bomba_generica.png como 1ª")
            continue
        try:
            if to_del:
                d = shopify_graphql(env, token, M_DEL_MEDIA, {"productId": p["id"], "mediaIds": to_del})
                if d["productDeleteMedia"]["mediaUserErrors"]:
                    raise RuntimeError(d["productDeleteMedia"]["mediaUserErrors"])
            res_url = staged_upload(env, token, "bomba_generica.png", GENERICA.read_bytes())
            alt = "Bomba hidráulica — imagem meramente ilustrativa | APP Agro Peças Padrão"
            d = shopify_graphql(env, token, M_CREATE_MEDIA, {
                "productId": p["id"],
                "media": [{"mediaContentType": "IMAGE", "originalSource": res_url, "alt": alt}]})
            r = d["productCreateMedia"]
            if r["mediaUserErrors"]:
                raise RuntimeError(r["mediaUserErrors"])
            mid = r["media"][0]["id"]
            if wait_ready(env, token, p["id"], mid):
                shopify_graphql(env, token, M_REORDER, {"id": p["id"], "moves": [{"id": mid, "newPosition": "0"}]})
            print(f"  ✔ {sku}: bomba_generica no lugar ({len(to_del)} mídia(s) removida(s))")
        except Exception as e:
            print(f"  ✖ {sku}: {e}")


if __name__ == "__main__":
    main()
