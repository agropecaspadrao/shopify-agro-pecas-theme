#!/usr/bin/env python3
"""
Gera planilhas mestras de preços e fretes para APP Agro Peças Padrão.

Saída (em util/):
  planilha_precos_master.xlsx   — formação de preço por produto + exportação Shopify
  planilha_fretes_master.xlsx   — política de fretes por distribuidor e região

Uso:
  python3 gerar_planilhas_precos.py
"""

import csv, pathlib, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE_DIR  = pathlib.Path(__file__).parent
CSV_FILE  = BASE_DIR.parent / "master_products_import.csv"

# ── Paleta de cores ──────────────────────────────────────────────────────────
C_GREEN_DARK  = "1B4332"
C_GREEN_MID   = "2F6B4F"
C_GREEN_LIGHT = "D6EAD8"
C_GOLD        = "D4AF37"
C_GOLD_LIGHT  = "FFF8DC"
C_ORANGE      = "E67E22"
C_ORANGE_LIGHT= "FDEBD0"
C_BLUE        = "2E86C1"
C_BLUE_LIGHT  = "D6EAF8"
C_GRAY        = "F0F0F0"
C_WHITE       = "FFFFFF"
C_RED_LIGHT   = "FDECEA"
C_HEADER_TEXT = "FFFFFF"

def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def fmt_brl(ws, row, col):
    ws.cell(row, col).number_format = 'R$ #,##0.00'
def fmt_pct(ws, row, col):
    ws.cell(row, col).number_format = '0.0%'

def style_header_dark(cell, text=None, size=10):
    if text: cell.value = text
    cell.fill = fill(C_GREEN_DARK)
    cell.font = font(bold=True, color=C_HEADER_TEXT, size=size)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border_thin()

def style_header_gold(cell, text=None):
    if text: cell.value = text
    cell.fill = fill(C_GOLD)
    cell.font = font(bold=True, color="000000", size=9)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border_thin()

def style_header_blue(cell, text=None):
    if text: cell.value = text
    cell.fill = fill(C_BLUE)
    cell.font = font(bold=True, color=C_HEADER_TEXT, size=9)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border_thin()

def style_header_orange(cell, text=None):
    if text: cell.value = text
    cell.fill = fill(C_ORANGE)
    cell.font = font(bold=True, color=C_HEADER_TEXT, size=9)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border_thin()

def style_input(cell):
    cell.fill = fill(C_GOLD_LIGHT)
    cell.font = font(bold=True, color="7B4F00", size=10)
    cell.alignment = align("center", "center")
    cell.border = border_thin()

def style_formula(cell):
    cell.fill = fill(C_GREEN_LIGHT)
    cell.font = font(color="1A4A2E", size=10)
    cell.alignment = align("center", "center")
    cell.border = border_thin()

def style_info(cell, wrap=False):
    cell.fill = fill(C_GRAY)
    cell.font = font(size=9)
    cell.alignment = align("left", "center", wrap=wrap)
    cell.border = border_thin()

def style_section(ws, row, col_start, col_end, text, color=C_GREEN_MID):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row, col_start, text)
    c.fill = fill(color)
    c.font = font(bold=True, color=C_HEADER_TEXT, size=10)
    c.alignment = align("center", "center")
    c.border = border_thin()

# ── Leitura do CSV ────────────────────────────────────────────────────────────
def load_products():
    products = []
    with open(CSV_FILE, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if not row["Handle"] or not row["Title"]:
                continue
            products.append({
                "handle": row["Handle"],
                "title":  row["Title"],
                "sku":    row["Variant SKU"],
                "type":   row["Type"],
                "tags":   row["Tags"],
            })
    return products

# ── Pesos estimados por tipo (kg/unidade) ────────────────────────────────────
PESO_TIPO = {
    "Bombas Hidráulicas":    6.5,
    "Sensores Agrícolas":    0.3,
    "GPS Agrícola":          0.8,
    "Monitores Agrícolas":   1.2,
    "Iluminação Agrícola":   0.5,
    "Kits Agrícolas":        0.4,
    "Peças para Plantadeira":0.2,
    "Peças para Trator":     0.3,
}
# Unidades mínimas para consolidação de frete
CONSOLIDACAO_TIPO = {
    "Bombas Hidráulicas":    1,
    "Sensores Agrícolas":    5,
    "GPS Agrícola":          1,
    "Monitores Agrícolas":   1,
    "Iluminação Agrícola":   3,
    "Kits Agrícolas":        3,
    "Peças para Plantadeira":20,
    "Peças para Trator":     10,
}

# ── Distribuidor por tipo (mapeamento padrão) ────────────────────────────────
DISTRIBUIDOR_TIPO = {
    "Bombas Hidráulicas":    "LIVENZA / Sohipren",
    "Sensores Agrícolas":    "Greco Agro Tech",
    "GPS Agrícola":          "Greco Agro Tech",
    "Monitores Agrícolas":   "Greco Agro Tech",
    "Iluminação Agrícola":   "Greco Agro Tech",
    "Kits Agrícolas":        "Greco Agro Tech",
    "Peças para Plantadeira":"AGCO (ADG)",
    "Peças para Trator":     "AGCO (ADG)",
}


# ════════════════════════════════════════════════════════════════════════════════
#  PLANILHA 1 — PREÇOS MASTER
# ════════════════════════════════════════════════════════════════════════════════

def criar_planilha_precos(products):
    wb = Workbook()

    # ── Aba 1: Configurações ─────────────────────────────────────────────────
    ws_cfg = wb.active
    ws_cfg.title = "⚙ Configurações"
    _aba_configuracoes(ws_cfg)

    # ── Aba 2: Formação de Preço ─────────────────────────────────────────────
    ws_fp = wb.create_sheet("💰 Formação de Preço")
    _aba_formacao_preco(ws_fp, products)

    # ── Aba 3: Exportação Shopify ────────────────────────────────────────────
    ws_ex = wb.create_sheet("🛒 Exportação Shopify")
    _aba_exportacao_shopify(ws_ex, products)

    out = BASE_DIR / "planilha_precos_master.xlsx"
    wb.save(out)
    print(f"✅ Planilha de preços salva em {out}")


def _aba_configuracoes(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55

    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "⚙  CONFIGURAÇÕES GLOBAIS — APP Agro Peças Padrão"
    c.fill = fill(C_GREEN_DARK); c.font = font(bold=True, color=C_HEADER_TEXT, size=14)
    c.alignment = align("center", "center")

    def bloco(row, titulo, cor=C_GREEN_MID):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, titulo)
        c.fill = fill(cor); c.font = font(bold=True, color=C_HEADER_TEXT, size=10)
        c.alignment = align("left", "center"); c.border = border_thin()

    def cfg_row(row, nome, valor, obs=""):
        ws.cell(row, 1, nome).border = border_thin()
        ws.cell(row, 1).font = font(size=9)
        ws.cell(row, 1).fill = fill(C_GRAY)
        ws.cell(row, 1).alignment = align("left", "center")
        c = ws.cell(row, 2, valor)
        style_input(c)
        ws.cell(row, 3, obs).border = border_thin()
        ws.cell(row, 3).font = font(size=8, italic=True, color="666666")
        ws.cell(row, 3).alignment = align("left", "center", wrap=True)

    # ── Custos Operacionais
    bloco(3, "CUSTOS OPERACIONAIS (% sobre Custo de Aquisição)")
    cfg_row(4,  "Overhead operacional (aluguel, salários, energia)",   0.18, "18% sobre o custo de aquisição — ajuste conforme DRE")
    cfg_row(5,  "Embalagem e manuseio",                                0.03, "3% sobre CA — inclui caixas, fita, etiqueta, NF")
    cfg_row(6,  "Comissão / marketplace / taxa Shopify",               0.025,"2.5% — taxa de gateway + plataforma")
    cfg_row(7,  "Impostos sobre venda (Simples Nacional)",             0.06, "6% médio Simples Nacional — verifique sua faixa")
    cfg_row(8,  "Inadimplência / perdas estimadas",                    0.01, "1% — provisão para devoluções e perdas")

    # ── Margem de Lucro
    bloco(10, "MARGENS DE LUCRO SUGERIDAS POR CATEGORIA")
    cfg_row(11, "Bombas Hidráulicas — margem mínima sobre venda",      0.35, "35% — produtos de alto valor, menor giro")
    cfg_row(12, "Sensores / GPS / Monitores — margem mínima",          0.40, "40% — produtos tech, margem maior")
    cfg_row(13, "Peças Plásticas / Plantadeira — margem mínima",       0.45, "45% — alto giro, margens maiores")
    cfg_row(14, "Margem de segurança adicional (concorrência)",         0.05, "5% extra para absorver variações de câmbio e insumos")

    # ── Frete Fábrica → APP
    bloco(16, "FRETE FÁBRICA → APP (Curitiba-PR)", C_BLUE)
    cfg_row(17, "LIVENZA/Sohipren (origem: SP ou importado)",          0.05, "~5% sobre PF — frete CIF/FOB Curitiba, peso médio 7kg")
    cfg_row(18, "Greco Agro Tech (origem: SP ou PR)",                  0.04, "~4% sobre PF — peças leves, frete mais barato")
    cfg_row(19, "AGCO / ADG (origem: SP)",                             0.045,"~4.5% sobre PF — peças médias/grandes")
    cfg_row(20, "ICMS interestadual médio (SP→PR)",                    0.07, "7% — verifique DIFAL e NCM de cada produto")

    # ── Frete APP → Cliente
    bloco(22, "FRETE APP → CLIENTE — BASE POR REGIÃO (R$/pedido)", C_ORANGE)
    cfg_row(23, "Sul próximo (PR) — base até 3kg",                     25.0, "Transportadora regional: Jamef, Braspress, TNT")
    cfg_row(24, "Sul (SC, RS) — base até 3kg",                         38.0, "")
    cfg_row(25, "Sudeste (SP, MG, RJ, ES) — base até 3kg",             48.0, "")
    cfg_row(26, "Centro-Oeste (MT, MS, GO, DF) — base até 3kg",        62.0, "")
    cfg_row(27, "Nordeste (BA, SE, AL, PE, PB, RN, CE, PI, MA) — base", 85.0,"")
    cfg_row(28, "Norte (PA, AM, TO, RO, AC, RR, AP) — base até 3kg", 110.0, "Correios ou transportadora regional")
    cfg_row(29, "Acréscimo por kg adicional acima de 3kg (R$/kg)",     14.0, "Média nacional para produtos industriais")
    cfg_row(30, "Seguro mínimo (R$)",                                   10.0, "Valor mínimo de seguro por envio")
    cfg_row(31, "Seguro % sobre NF",                                   0.004,"0,4% do valor declarado — incluso nas transportadoras sérias")

    for row in range(3, 32):
        ws.row_dimensions[row].height = 22

    # ── nota
    ws.merge_cells("A33:C36")
    nota = ws["A33"]
    nota.value = (
        "📌  COMO USAR:\n"
        "1. Ajuste os valores em amarelo conforme sua realidade financeira.\n"
        "2. Na aba '💰 Formação de Preço', preencha APENAS as colunas amarelas (Preço de Fábrica, Markup desejado, Preço mín./máx. mercado).\n"
        "3. Todas as colunas verdes são calculadas automaticamente por fórmula.\n"
        "4. Na aba '🛒 Exportação Shopify', copie e cole no arquivo master_products_import.csv para subir preços em lote."
    )
    nota.fill = fill("FFFDE7"); nota.font = font(size=9, italic=True)
    nota.alignment = align("left", "top", wrap=True)
    nota.border = border_medium()
    ws.row_dimensions[33].height = 80


def _aba_formacao_preco(ws, products):
    ws.sheet_view.showGridLines = False

    # ── cabeçalho título ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:AB1")
    c = ws["A1"]
    c.value = "💰  TABELA MESTRA DE FORMAÇÃO DE PREÇO — APP Agro Peças Padrão"
    c.fill = fill(C_GREEN_DARK); c.font = font(bold=True, color=C_HEADER_TEXT, size=13)
    c.alignment = align("center", "center")

    # ── grupos de colunas ────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    groups = [
        (1,  5,  "IDENTIFICAÇÃO DO PRODUTO",            C_GREEN_DARK),
        (6,  9,  "CUSTO DE AQUISIÇÃO (fábrica → APP)",  C_BLUE),
        (10, 13, "CUSTOS OPERACIONAIS",                  C_GREEN_MID),
        (14, 16, "CUSTO TOTAL & MARKUP",                 "5D4037"),  # marrom
        (17, 22, "FRETE APP → CLIENTE (por região)",     C_ORANGE),
        (23, 25, "SEGURO & PREÇO FINAL",                 C_GREEN_MID),
        (26, 27, "PESQUISA DE MERCADO",                  "6A1B9A"),  # roxo
        (28, 28, "EXPORTAÇÃO SHOPIFY",                   "C62828"),  # vermelho
    ]
    for cs, ce, txt, cor in groups:
        style_section(ws, 2, cs, ce, txt, cor)

    # ── cabeçalhos de coluna ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 48
    headers = [
        # Identificação
        ("A", "Handle\n(Shopify)",        14),
        ("B", "Título",                   38),
        ("C", "SKU",                      22),
        ("D", "Distribuidor",             22),
        ("E", "Tipo / Categoria",         20),
        # Custo de Aquisição
        ("F", "⭐ Preço\nde Fábrica\n(R$)", 14),
        ("G", "% Frete\nFábrica→APP",      12),
        ("H", "Frete\nFábrica→APP\n(R$)",  13),
        ("I", "ICMS\nEntrada\n(R$)",       12),
        # Custos Operacionais
        ("J", "% Overhead\n+ Embalagem",   12),
        ("K", "Overhead\n+ Embalagem\n(R$)",13),
        ("L", "% Impostos\n+ Comissão",    12),
        ("M", "Impostos\n+ Comissão\n(R$)",13),
        # Custo Total
        ("N", "Custo Total\n(R$)",         14),
        ("O", "⭐ Markup\nDesejado %",     13),
        ("P", "Preço de\nVenda s/Frete\n(R$)", 14),
        # Frete por região
        ("Q", "Frete Sul\nPR (R$)",        11),
        ("R", "Frete Sul\nSC/RS (R$)",     11),
        ("S", "Frete\nSudeste (R$)",       11),
        ("T", "Frete\nC-Oeste (R$)",       11),
        ("U", "Frete\nNordeste (R$)",      12),
        ("V", "Frete Norte\n(R$)",         11),
        # Seguro & Final
        ("W", "Seguro\nEstimado (R$)",     12),
        ("X", "Peso Unit.\n(kg)",          10),
        ("Y", "Margem Real\n% s/ Venda",   12),
        # Mercado
        ("Z", "⭐ Preço Mín.\nMercado (R$)",14),
        ("AA","⭐ Preço Máx.\nMercado (R$)",14),
        # Shopify
        ("AB","Preço Shopify\n(Variant Price)",15),
    ]

    for col_letter, header, width in headers:
        col_idx = ord(col_letter) - ord('A') + 1 if len(col_letter) == 1 else \
                  (ord(col_letter[0]) - ord('A') + 1) * 26 + (ord(col_letter[1]) - ord('A') + 1)
        c = ws.cell(3, col_idx, header)
        # Cor por grupo
        if col_idx <= 5:   style_header_dark(c)
        elif col_idx <= 9: style_header_blue(c)
        elif col_idx <= 13:
            c.fill = fill(C_GREEN_MID); c.font = font(bold=True, color=C_HEADER_TEXT, size=9)
            c.alignment = align("center","center",wrap=True); c.border = border_thin()
        elif col_idx <= 16:
            c.fill = fill("5D4037"); c.font = font(bold=True, color=C_HEADER_TEXT, size=9)
            c.alignment = align("center","center",wrap=True); c.border = border_thin()
        elif col_idx <= 22: style_header_orange(c)
        elif col_idx <= 25:
            c.fill = fill(C_GREEN_MID); c.font = font(bold=True, color=C_HEADER_TEXT, size=9)
            c.alignment = align("center","center",wrap=True); c.border = border_thin()
        elif col_idx <= 27:
            c.fill = fill("6A1B9A"); c.font = font(bold=True, color=C_HEADER_TEXT, size=9)
            c.alignment = align("center","center",wrap=True); c.border = border_thin()
        else:
            c.fill = fill("C62828"); c.font = font(bold=True, color=C_HEADER_TEXT, size=9)
            c.alignment = align("center","center",wrap=True); c.border = border_thin()

        ws.column_dimensions[col_letter if len(col_letter)==1 else col_letter].width = width

    ws.column_dimensions["AA"].width = 14
    ws.column_dimensions["AB"].width = 15

    # ── Linhas de produto ────────────────────────────────────────────────────
    cfg = "⚙ Configurações"   # nome da aba de config

    # Índices de linha das configurações (na aba Config):
    # B4=overhead, B5=embalagem, B6=comissão, B7=impostos, B8=inadimpl.
    # B11=margem bombas, B12=margem sensores, B13=margem plasticos
    # B17=frete Livenza, B18=frete Greco, B19=frete AGCO, B20=ICMS
    # B23..B28=frete regiões, B29=taxa/kg, B30=seg min, B31=seg%

    for i, p in enumerate(products):
        r = i + 4
        ws.row_dimensions[r].height = 20

        distribuidor = DISTRIBUIDOR_TIPO.get(p["type"], "—")
        peso         = PESO_TIPO.get(p["type"], 1.0)

        # Frete fábrica % (col na config)
        if "LIVENZA" in distribuidor or "Sohipren" in distribuidor:
            pct_frete_fab = f"'⚙ Configurações'!$B$17"
        elif "Greco" in distribuidor:
            pct_frete_fab = f"'⚙ Configurações'!$B$18"
        else:
            pct_frete_fab = f"'⚙ Configurações'!$B$19"

        # Margem padrão por tipo
        if "Bomba" in p["type"]:
            pct_markup_default = f"'⚙ Configurações'!$B$11"
        elif any(x in p["type"] for x in ["Sensor","GPS","Monitor","Iluminação","Kit"]):
            pct_markup_default = f"'⚙ Configurações'!$B$12"
        else:
            pct_markup_default = f"'⚙ Configurações'!$B$13"

        def cell(col):
            return f"{get_column_letter(col)}{r}"

        # Colunas 1-5: info
        ws.cell(r, 1, p["handle"]); style_info(ws.cell(r,1))
        ws.cell(r, 2, p["title"]);  ws.cell(r,2).fill = fill(C_GRAY); ws.cell(r,2).font = font(size=9); ws.cell(r,2).border = border_thin(); ws.cell(r,2).alignment = align("left","center",wrap=True)
        ws.cell(r, 3, p["sku"]);    style_info(ws.cell(r,3)); ws.cell(r,3).alignment=align("center","center")
        ws.cell(r, 4, distribuidor);style_info(ws.cell(r,4))
        ws.cell(r, 5, p["type"]);   style_info(ws.cell(r,5))

        # Col 6: Preço de Fábrica — INPUT DO USUÁRIO
        ws.cell(r, 6, 0.0); style_input(ws.cell(r,6)); ws.cell(r,6).number_format = 'R$ #,##0.00'

        # Col 7: % frete fábrica (de config)
        ws.cell(r, 7, f"={pct_frete_fab}"); style_formula(ws.cell(r,7)); fmt_pct(ws,r,7)

        # Col 8: Frete fábrica→APP = F*G
        ws.cell(r, 8, f"=F{r}*G{r}"); style_formula(ws.cell(r,8)); ws.cell(r,8).number_format='R$ #,##0.00'

        # Col 9: ICMS = F * config!B20
        ws.cell(r, 9, f"=F{r}*'⚙ Configurações'!$B$20"); style_formula(ws.cell(r,9)); ws.cell(r,9).number_format='R$ #,##0.00'

        # Col 10: % overhead+embalagem = config!B4+config!B5
        ws.cell(r,10, f"='⚙ Configurações'!$B$4+'⚙ Configurações'!$B$5"); style_formula(ws.cell(r,10)); fmt_pct(ws,r,10)

        # Col 11: R$ overhead = (F+H+I)*J
        ws.cell(r,11, f"=(F{r}+H{r}+I{r})*J{r}"); style_formula(ws.cell(r,11)); ws.cell(r,11).number_format='R$ #,##0.00'

        # Col 12: % impostos+comissão = config!B6+config!B7+config!B8
        ws.cell(r,12, f"='⚙ Configurações'!$B$6+'⚙ Configurações'!$B$7+'⚙ Configurações'!$B$8"); style_formula(ws.cell(r,12)); fmt_pct(ws,r,12)

        # Col 13: R$ impostos = sobre preço de venda (calculado iterativo simplificado: custo_base * L/(1-L))
        ws.cell(r,13, f"=N{r}*L{r}/(1-L{r})"); style_formula(ws.cell(r,13)); ws.cell(r,13).number_format='R$ #,##0.00'

        # Col 14: Custo Total = F+H+I+K+M
        ws.cell(r,14, f"=F{r}+H{r}+I{r}+K{r}+M{r}"); style_formula(ws.cell(r,14))
        ws.cell(r,14).font = font(bold=True, color="1A4A2E"); ws.cell(r,14).number_format='R$ #,##0.00'
        ws.cell(r,14).fill = fill("B8E0BF")

        # Col 15: Markup % — INPUT (pré-preenchido com padrão da config)
        ws.cell(r,15, f"={pct_markup_default}"); style_input(ws.cell(r,15)); fmt_pct(ws,r,15)

        # Col 16: Preço de Venda s/Frete = N/(1-O)
        ws.cell(r,16, f"=IF(O{r}>=1,N{r}*2,N{r}/(1-O{r}))"); style_formula(ws.cell(r,16))
        ws.cell(r,16).font = font(bold=True, color="1A4A2E"); ws.cell(r,16).fill = fill("B8E0BF")
        ws.cell(r,16).number_format='R$ #,##0.00'

        # Colunas 17-22: Frete por região = base_config + max(0, (peso-3)*taxa/kg)
        regioes_cfg = [23,24,25,26,27,28]
        for j, cfg_row_num in enumerate(regioes_cfg):
            col = 17 + j
            formula = (f"=IF(F{r}=0,0,"
                       f"ROUND('⚙ Configurações'!$B${cfg_row_num}"
                       f"+MAX(0,({peso}-3)*'⚙ Configurações'!$B$29),2))")
            ws.cell(r, col, formula)
            ws.cell(r,col).fill = fill(C_ORANGE_LIGHT)
            ws.cell(r,col).font = font(size=9, color="7B3A00")
            ws.cell(r,col).number_format='R$ #,##0.00'
            ws.cell(r,col).border = border_thin()
            ws.cell(r,col).alignment = align("center","center")

        # Col 23: Seguro = MAX(config!B30, P*config!B31)
        ws.cell(r,23, f"=IF(F{r}=0,0,ROUND(MAX('⚙ Configurações'!$B$30,P{r}*'⚙ Configurações'!$B$31),2))")
        style_formula(ws.cell(r,23)); ws.cell(r,23).number_format='R$ #,##0.00'

        # Col 24: Peso unitário (referência)
        ws.cell(r,24, peso); style_info(ws.cell(r,24)); ws.cell(r,24).alignment=align("center","center")
        ws.cell(r,24).number_format='0.0 "kg"'

        # Col 25: Margem real % = (P-N)/P
        ws.cell(r,25, f"=IF(P{r}=0,0,(P{r}-N{r})/P{r})"); style_formula(ws.cell(r,25)); fmt_pct(ws,r,25)
        ws.cell(r,25).font = font(bold=True, color="1A4A2E")

        # Col 26: Preço mín mercado — INPUT
        ws.cell(r,26, 0.0); style_input(ws.cell(r,26)); ws.cell(r,26).fill=fill("F3E5F5")
        ws.cell(r,26).font = font(bold=True, color="4A148C"); ws.cell(r,26).number_format='R$ #,##0.00'

        # Col 27: Preço máx mercado — INPUT
        ws.cell(r,27, 0.0); style_input(ws.cell(r,27)); ws.cell(r,27).fill=fill("F3E5F5")
        ws.cell(r,27).font = font(bold=True, color="4A148C"); ws.cell(r,27).number_format='R$ #,##0.00'

        # Col 28: Preço Shopify = P (preço de venda sem frete)
        ws.cell(r,28, f"=IF(P{r}>0,ROUND(P{r},2),\"\")")
        ws.cell(r,28).fill = fill("FFEBEE"); ws.cell(r,28).font = font(bold=True, color="B71C1C")
        ws.cell(r,28).number_format='R$ #,##0.00'; ws.cell(r,28).border = border_thin()
        ws.cell(r,28).alignment = align("center","center")

    # ── Freeze panes e filtros ───────────────────────────────────────────────
    ws.freeze_panes = "F4"
    ws.auto_filter.ref = f"A3:AB{3+len(products)}"

    # ── Legenda ──────────────────────────────────────────────────────────────
    leg_row = len(products) + 5
    ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=28)
    c = ws.cell(leg_row, 1)
    c.value = ("⭐ Colunas com estrela = preencher manualmente     "
               "🟡 Amarelo = input     🟢 Verde = calculado automaticamente     "
               "🟠 Laranja = frete por região (automático)")
    c.fill = fill("FFFDE7"); c.font = font(size=9, italic=True); c.border = border_medium()
    c.alignment = align("left","center")


def _aba_exportacao_shopify(ws, products):
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "🛒  EXPORTAÇÃO SHOPIFY — Cole estes dados no Shopify Admin → Products → Import"
    c.fill = fill("C62828"); c.font = font(bold=True, color=C_HEADER_TEXT, size=12)
    c.alignment = align("center","center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "Copie a coluna 'Variant Price' e cole no campo correspondente do master_products_import.csv para subir preços em lote."
    c.fill = fill("FFEBEE"); c.font = font(size=9, italic=True, color="B71C1C")
    c.alignment = align("center","center")

    headers_ex = ["Handle", "Título (ref.)", "Variant Price (R$)", "Variant Compare At Price (R$)"]
    widths_ex   = [36, 52, 22, 26]
    for j, (h, w) in enumerate(zip(headers_ex, widths_ex)):
        col = j+1
        c = ws.cell(3, col, h)
        style_header_dark(c); ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[3].height = 24
    fp_sheet = "💰 Formação de Preço"
    for i, p in enumerate(products):
        r = i + 4
        ws.cell(r, 1, p["handle"]); style_info(ws.cell(r,1)); ws.cell(r,1).alignment=align("left","center")
        ws.cell(r, 2, p["title"]);  style_info(ws.cell(r,2)); ws.cell(r,2).alignment=align("left","center",wrap=True)
        # Puxa Preço de Venda da aba de formação (coluna P = 16)
        prod_row = i + 4
        ws.cell(r, 3, f"=IF('💰 Formação de Preço'!P{prod_row}>0,ROUND('💰 Formação de Preço'!P{prod_row},2),\"\")")
        ws.cell(r,3).fill = fill("FFEBEE"); ws.cell(r,3).font = font(bold=True, color="B71C1C")
        ws.cell(r,3).number_format='0.00'; ws.cell(r,3).border = border_thin(); ws.cell(r,3).alignment=align("center","center")
        # Compare At Price = Preço Máx mercado (col AA=27)
        ws.cell(r, 4, f"=IF('💰 Formação de Preço'!AA{prod_row}>0,ROUND('💰 Formação de Preço'!AA{prod_row},2),\"\")")
        ws.cell(r,4).fill = fill("F3E5F5"); ws.cell(r,4).font = font(bold=True, color="4A148C")
        ws.cell(r,4).number_format='0.00'; ws.cell(r,4).border = border_thin(); ws.cell(r,4).alignment=align("center","center")
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"
    nota_r = len(products)+5
    ws.merge_cells(start_row=nota_r, start_column=1, end_row=nota_r+2, end_column=4)
    c = ws.cell(nota_r,1)
    c.value = (
        "📤 COMO EXPORTAR PARA O SHOPIFY:\n"
        "1. Abra o arquivo master_products_import.csv em editor de texto ou Excel.\n"
        "2. Substitua os valores na coluna 'Variant Price' pelos da coluna C desta aba.\n"
        "3. Importe no Shopify Admin → Products → Import → 'Overwrite existing products'.\n"
        "4. O campo 'Variant Compare At Price' exibe o preço riscado (pesquisa de mercado máximo)."
    )
    c.fill = fill("FFFDE7"); c.font = font(size=9, italic=True)
    c.alignment = align("left","top",wrap=True); c.border = border_medium()
    for rr in range(nota_r, nota_r+3):
        ws.row_dimensions[rr].height = 24


# ════════════════════════════════════════════════════════════════════════════════
#  PLANILHA 2 — FRETES MASTER
# ════════════════════════════════════════════════════════════════════════════════

def criar_planilha_fretes():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "🏭 Fábrica → APP"
    _aba_frete_fabrica(ws1)

    ws2 = wb.create_sheet("🚚 APP → Cliente")
    _aba_frete_cliente(ws2)

    ws3 = wb.create_sheet("📦 Embalagem Progressiva")
    _aba_embalagem_progressiva(ws3)

    ws4 = wb.create_sheet("🛡 Seguro de Carga")
    _aba_seguro(ws4)

    ws5 = wb.create_sheet("🗺 Regiões do Agro BR")
    _aba_regioes(ws5)

    out = BASE_DIR / "planilha_fretes_master.xlsx"
    wb.save(out)
    print(f"✅ Planilha de fretes salva em {out}")


def _titulo(ws, texto, col_end=10):
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f"A1:{get_column_letter(col_end)}1")
    c = ws["A1"]
    c.value = texto; c.fill = fill(C_GREEN_DARK)
    c.font = font(bold=True, color=C_HEADER_TEXT, size=13)
    c.alignment = align("center","center")
    ws.sheet_view.showGridLines = False


def _aba_frete_fabrica(ws):
    _titulo(ws, "🏭  FRETE FÁBRICA → APP (Centro de Distribuição Curitiba-PR)", 8)

    ws.row_dimensions[3].height = 30
    headers = ["Distribuidor","Origem (Cidade/Estado)","Modal","Prazo Estimado","Peso Médio Lote (kg)",
               "% Frete s/ Valor NF","Frete Mín (R$)","Observações"]
    widths  = [28, 30, 18, 18, 20, 18, 16, 45]
    for j,(h,w) in enumerate(zip(headers,widths)):
        c = ws.cell(3, j+1, h); style_header_dark(c)
        ws.column_dimensions[get_column_letter(j+1)].width = w

    dados = [
        ["LIVENZA / Sohipren", "São Paulo - SP", "Transportadora (CIF)", "3–5 dias úteis", 350,  0.050, 80.0,  "Pedidos acima de R$5.000 geralmente CIF. Verificar incoterm no pedido."],
        ["Greco Agro Tech",    "Curitiba - PR (ou SP)", "Própria / Correios", "1–2 dias úteis", 50,  0.030, 20.0,  "Fabricante próximo. Frete pode ser isento acima de certo valor mínimo de pedido."],
        ["AGCO / ADG",         "São Paulo - SP", "Transportadora (FOB)", "4–6 dias úteis", 120, 0.045, 55.0,  "Frete FOB — APP paga. Agrupar pedidos para reduzir custo unitário."],
    ]

    for i, row_data in enumerate(dados):
        r = i + 4
        ws.row_dimensions[r].height = 24
        for j, val in enumerate(row_data):
            c = ws.cell(r, j+1, val)
            c.border = border_thin(); c.alignment = align("center","center",wrap=(j in (0,7)))
            c.font = font(size=9)
            if j == 0: c.fill = fill(C_GREEN_LIGHT); c.font = font(bold=True, size=9, color="1A4A2E")
            elif j == 5: c.number_format = '0.0%'; c.fill = fill(C_GOLD_LIGHT); c.font=font(bold=True,size=9,color="7B4F00")
            elif j == 6: c.number_format = 'R$ #,##0.00'; c.fill = fill(C_GOLD_LIGHT); c.font=font(bold=True,size=9,color="7B4F00")
            else: c.fill = fill(C_GRAY if r%2==0 else C_WHITE)
            if j == 7: c.alignment = align("left","center",wrap=True)

    # nota
    nota_r = 9
    ws.row_dimensions[nota_r].height = 60
    ws.merge_cells(f"A{nota_r}:H{nota_r}")
    c = ws.cell(nota_r,1)
    c.value = ("⚠  IMPORTANTE: Os percentuais de frete acima são médias de mercado para referência. "
               "Solicite tabelas atualizadas das transportadoras (Jamef, Braspress, TNT, Rodonaves) periodicamente. "
               "Para pedidos grandes (>R$10.000), negocie frete CIF diretamente com o distribuidor. "
               "ICMS interestadual (SP→PR): 7% para produtos industriais. Verifique NCM e DIFAL.")
    c.fill = fill("FFF8DC"); c.font = font(size=9, italic=True)
    c.alignment = align("left","center",wrap=True); c.border = border_medium()


def _aba_frete_cliente(ws):
    _titulo(ws, "🚚  FRETE APP (Curitiba-PR) → CLIENTE FINAL — Tabela por Região e Peso", 10)

    # sub-título
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "Tarifas de referência — Transportadoras regionais (Jamef, Braspress, TNT, Rodonaves) + Correios. Atualize semestralmente."
    c.fill = fill("E8F5E9"); c.font = font(size=9, italic=True, color="1A4A2E")
    c.alignment = align("center","center")
    ws.row_dimensions[2].height = 20

    # cabeçalho pesos
    ws.row_dimensions[3].height = 24
    ws.cell(3,1,"Região"); style_header_dark(ws.cell(3,1)); ws.column_dimensions["A"].width=32
    ws.cell(3,2,"Estados Inclusos"); style_header_dark(ws.cell(3,2)); ws.column_dimensions["B"].width=50
    pesos = ["Até 1 kg","1–3 kg","3–7 kg","7–15 kg","15–30 kg","Acima 30kg","Prazo (d.u.)","Transportadora ref."]
    widths_p = [12,12,12,12,13,13,13,28]
    for j,(h,w) in enumerate(zip(pesos,widths_p)):
        c = ws.cell(3, j+3, h); style_header_gold(c)
        ws.column_dimensions[get_column_letter(j+3)].width = w

    regioes = [
        ["Sul Próximo — Paraná (PR)",         "Curitiba, Londrina, Maringá, Cascavel, Foz",
         22, 30, 42, 62, 95, 140, "1–2", "Transportadora regional / próprio"],
        ["Sul — Santa Catarina (SC)",          "Florianópolis, Joinville, Blumenau, Chapecó",
         30, 40, 56, 82, 125, 185, "2–3", "Jamef / Braspress"],
        ["Sul — Rio Grande do Sul (RS)",       "Porto Alegre, Caxias do Sul, Pelotas, Santa Maria",
         35, 48, 65, 95, 145, 215, "2–3", "Jamef / TNT"],
        ["Sudeste — São Paulo (SP)",           "São Paulo, Campinas, Ribeirão Preto, Santos",
         40, 52, 72, 105, 160, 240, "2–3", "Braspress / TNT"],
        ["Sudeste — Minas Gerais (MG)",        "Belo Horizonte, Uberlândia, Juiz de Fora",
         44, 58, 80, 118, 175, 260, "3–4", "Braspress / Jamef"],
        ["Sudeste — RJ / ES",                  "Rio de Janeiro, Vitória, Campos",
         46, 62, 85, 125, 185, 275, "3–5", "TNT / Braspress"],
        ["Centro-Oeste — MS / GO / DF",        "Campo Grande, Goiânia, Brasília, Dourados",
         52, 70, 96, 140, 210, 310, "3–5", "Rodonaves / Jamef"],
        ["Centro-Oeste — Mato Grosso (MT)",    "Cuiabá, Rondonópolis, Sorriso, Lucas do Rio Verde",
         60, 82, 112, 165, 245, 365, "4–6", "Rodonaves / J.Leite"],
        ["Nordeste — Bahia (BA) / Sergipe (SE)","Salvador, Feira de Santana, Aracaju",
         70, 95, 130, 190, 280, 415, "5–7", "TNT / Correios Sedex Empresarial"],
        ["Nordeste — PE / AL / PB / RN",       "Recife, Maceió, João Pessoa, Natal",
         78, 105, 142, 208, 308, 455, "5–8", "TNT / Correios"],
        ["Nordeste — CE / PI / MA",            "Fortaleza, Teresina, São Luís",
         85, 115, 156, 228, 338, 500, "6–9", "Correios / Golog"],
        ["Norte — Pará (PA) / Tocantins (TO)", "Belém, Santarém, Palmas",
         95, 128, 175, 255, 378, 560, "7–10","Correios Sedex / transportadora"],
        ["Norte — AM / RO / AC / RR / AP",     "Manaus, Porto Velho, Rio Branco, Macapá",
         115, 155, 210, 308, 455, 675, "8–14","Correios (único modal viável em vários municípios)"],
    ]

    for i, reg in enumerate(regioes):
        r = i + 4
        ws.row_dimensions[r].height = 22
        fundo = C_GRAY if i%2==0 else C_WHITE
        for j, val in enumerate(reg):
            c = ws.cell(r, j+1, val)
            c.border = border_thin()
            c.font = font(size=9, bold=(j==0))
            c.fill = fill(C_GREEN_LIGHT if j==0 else fundo)
            if j == 0: c.alignment=align("left","center",wrap=True)
            elif j == 1: c.alignment=align("left","center",wrap=True)
            elif 2<=j<=7: c.number_format='R$ #,##0.00'; c.alignment=align("center","center")
            else: c.alignment=align("center","center",wrap=True)

    # Tabela de frete progressivo (quantidade)
    ws.row_dimensions[19].height = 22
    ws.merge_cells("A19:J19")
    style_section(ws, 19, 1, 10, "📦  FRETE PROGRESSIVO — Desconto por Quantidade no Mesmo Pedido", C_ORANGE)

    desc_headers = ["Qtd unidades no pedido","Fator multiplicador no frete","Exemplo prático"]
    desc_widths   = [28, 24, 55]
    ws.row_dimensions[20].height = 24
    for j,(h,w) in enumerate(zip(desc_headers,desc_widths)):
        c = ws.cell(20, j+1, h); style_header_orange(c)

    descontos = [
        ["1 unidade",                  1.00, "Pedido normal — frete cheio"],
        ["2–5 unidades mesma linha",   0.85, "15% desc. — cabe em 1 caixa"],
        ["6–20 unidades mesma linha",  0.70, "30% desc. — 1 palete pequeno"],
        ["21–100 unidades",            0.55, "45% desc. — palete completo"],
        ["Acima de 100 unidades",      0.40, "60% desc. — frete consolidado/fracionado"],
    ]
    for i, row_d in enumerate(descontos):
        r = 21 + i
        ws.row_dimensions[r].height = 20
        for j, val in enumerate(row_d):
            c = ws.cell(r, j+1, val)
            c.border = border_thin(); c.font = font(size=9)
            c.fill = fill(C_ORANGE_LIGHT if i%2==0 else C_WHITE)
            c.alignment = align("center","center") if j<2 else align("left","center")
            if j==1: c.number_format='0%'; c.font=font(bold=True,size=9,color="7B3A00")


def _aba_embalagem_progressiva(ws):
    _titulo(ws, "📦  EMBALAGEM PROGRESSIVA — Consolidação de Volumes por Tipo de Produto", 8)

    ws.row_dimensions[2].height=18
    ws.merge_cells("A2:H2")
    c=ws["A2"]; c.value="A consolidação reduz o frete total ao enviar múltiplas unidades em uma única caixa/palete, diminuindo o custo por unidade."
    c.fill=fill("E8F5E9"); c.font=font(size=9,italic=True); c.alignment=align("center","center")

    ws.row_dimensions[3].height=28
    headers=["Categoria","Peso Unit. (kg)","Dimensões Unit. (cm)","Unid. por caixa","Caixa (kg)","Caixa (L×W×H cm)","Peso volumétrico (kg)","Observações"]
    widths= [28,14,22,14,12,22,18,45]
    for j,(h,w) in enumerate(zip(headers,widths)):
        c=ws.cell(3,j+1,h); style_header_dark(c)
        ws.column_dimensions[get_column_letter(j+1)].width=w

    emb=[
        ["Bomba Hidráulica (pequena)",   2.5, "25×20×15",   1, 3.2,  "30×25×20",  "2,5",  "1 bomba/caixa — sempre individual"],
        ["Bomba Hidráulica (média)",     5.0, "30×25×20",   1, 6.0,  "35×30×25",  "4,4",  "1 bomba/caixa — caixa reforçada"],
        ["Bomba Hidráulica (dupla/grande)",9.0,"40×30×25",  1,10.5,  "45×35×30",  "7,1",  "Peso real > volumétrico — cobrar peso real"],
        ["Sensor Agrícola (pequeno)",    0.15,"10×8×5",    10, 1.8,  "25×20×12",  "1,0",  "10 sensores/caixa — embalagem Greco Agro Tech"],
        ["Sensor de Fluxo (médio)",      0.30,"15×10×8",    5, 1.8,  "22×18×12",  "0,7",  ""],
        ["GPS / Monitor Agrícola",       0.80,"20×15×8",    3, 2.8,  "32×22×12",  "1,4",  "Embalagem original do fabricante"],
        ["Kit Ponta de Cerca",           0.35,"20×15×5",    8, 3.2,  "30×22×16",  "1,7",  "Bolsas + caixa master"],
        ["Peças Plásticas (pequenas)",   0.05,"5×4×3",    100, 5.5,  "35×25×20",  "5,8",  "Sacos plásticos + caixa. Peso volumétrico ≈ real"],
        ["Peças Plásticas (médias)",     0.15,"10×8×5",    30, 4.8,  "32×28×20",  "6,0",  ""],
        ["Dobradiças / Acoplamentos",    0.30,"15×10×6",   20, 6.5,  "35×25×18",  "5,3",  "Caixa média — metal, peso real > volumétrico"],
    ]

    for i, row_e in enumerate(emb):
        r=i+4; ws.row_dimensions[r].height=20
        fundo=C_GRAY if i%2==0 else C_WHITE
        for j,val in enumerate(row_e):
            c=ws.cell(r,j+1,val); c.border=border_thin(); c.font=font(size=9)
            c.fill=fill(C_GREEN_LIGHT if j==0 else fundo)
            c.alignment=align("left","center") if j in (0,5,7) else align("center","center")
            if j in (1,4): c.number_format='0.0 "kg"'

    nota_r=15
    ws.row_dimensions[nota_r].height=55
    ws.merge_cells(f"A{nota_r}:H{nota_r}")
    c=ws.cell(nota_r,1)
    c.value=("📐 PESO VOLUMÉTRICO = (Comprimento × Largura × Altura em cm) ÷ 6.000\n"
             "As transportadoras cobram o MAIOR valor entre peso real e peso volumétrico.\n"
             "Exemplo: caixa 40×30×25cm = 30.000cm³ ÷ 6.000 = 5,0 kg volumétrico. Se a bomba pesa 6kg real → cobra 6kg.")
    c.fill=fill("FFF8DC"); c.font=font(size=9,italic=True); c.border=border_medium(); c.alignment=align("left","top",wrap=True)


def _aba_seguro(ws):
    _titulo(ws, "🛡  SEGURO DE CARGA — Cálculo e Política de Seguro para Envios", 7)

    ws.merge_cells("A2:G2")
    c=ws["A2"]; c.value="O seguro protege contra extravio, danos e roubo durante o transporte. A maioria das transportadoras inclui seguro básico até 10% do valor do frete."
    c.fill=fill("E8F5E9"); c.font=font(size=9,italic=True); c.alignment=align("center","center")

    ws.row_dimensions[3].height=26
    for j,(h,w) in enumerate(zip(
        ["Faixa de Valor do Produto (R$)","Seguro Mínimo (R$)","% Seguro","Seguro Calculado Exemplo","Cobertura Recomendada","Modal","Observações"],
        [35,18,14,28,30,20,50])):
        c=ws.cell(3,j+1,h); style_header_dark(c); ws.column_dimensions[get_column_letter(j+1)].width=w

    seguros=[
        ["Até R$ 500",        10.0, 0.004, "R$ 10,00 (mín)",      "Correios AR / Sedex Empresarial",    "Correios",       "Seguro incluso no Sedex — declarar valor correto"],
        ["R$ 500 – R$ 2.000", 10.0, 0.004, "R$ 8 a R$ 10 (mín)", "Seguro adicional transportadora",    "Transportadora","Solicitar ADE (Aviso de Despacho)"],
        ["R$ 2.000 – R$ 5.000",20.0,0.004, "R$ 20 a R$ 45",      "Seguro obrigatório + cobertura total","Transportadora","Exigir CTE com seguro declarado"],
        ["R$ 5.000 – R$ 15.000",50.0,0.004,"R$ 50 a R$ 200",     "Seguro full + rastreamento",          "Transportadora","Seguro avulso (~0,4-0,5% do valor)"],
        ["Acima de R$ 15.000",100.0,0.005, "R$ 150+",             "Seguro especializado de carga",       "Transportadora","Contratar apólice específica com corretora"],
    ]

    for i,row_s in enumerate(seguros):
        r=i+4; ws.row_dimensions[r].height=22
        fundo=C_GRAY if i%2==0 else C_WHITE
        for j,val in enumerate(row_s):
            c=ws.cell(r,j+1,val); c.border=border_thin(); c.font=font(size=9)
            c.fill=fill(C_BLUE_LIGHT if j==0 else fundo)
            c.alignment=align("left","center") if j in (0,4,5,6) else align("center","center")
            if j==1: c.number_format='R$ #,##0.00'
            if j==2: c.number_format='0.0%'

    ws.row_dimensions[10].height=22
    ws.merge_cells("A10:G10")
    style_section(ws,10,1,7,"💡 FÓRMULA: Seguro = MAX(valor_mínimo; valor_produto × 0,4%)",C_BLUE)

    nota_r=11
    ws.row_dimensions[nota_r].height=70
    ws.merge_cells(f"A{nota_r}:G{nota_r}")
    c=ws.cell(nota_r,1)
    c.value=("📌 BOAS PRÁTICAS:\n"
             "• Sempre declarar o valor real da NF — seguro calculado sobre valor declarado\n"
             "• Para bombas hidráulicas (R$800-3.000 tipicamente): seguro mínimo R$20, ideal R$35\n"
             "• Nunca subestimar o valor para pagar menos seguro — em caso de sinistro o reembolso será proporcional\n"
             "• Incluir custo do seguro no preço final (embutido) ou cobrar separado como 'proteção de envio'")
    c.fill=fill("E3F2FD"); c.font=font(size=9,italic=True); c.border=border_medium(); c.alignment=align("left","top",wrap=True)


def _aba_regioes(ws):
    _titulo(ws, "🗺  REGIÕES DO AGRONEGÓCIO BRASILEIRO — Mapa de Referência para Fretes", 5)

    ws.merge_cells("A2:E2")
    c=ws["A2"]; c.value="Referência para calcular fretes regionais. Distâncias a partir de Curitiba-PR (Centro de Distribuição APP)."
    c.fill=fill("E8F5E9"); c.font=font(size=9,italic=True); c.alignment=align("center","center")

    ws.row_dimensions[3].height=26
    for j,(h,w) in enumerate(zip(
        ["Região / Estado","Principais Polos do Agro","Distância de Curitiba (km)","Prazo Médio (d.u.)","Observações Logísticas"],
        [30,55,26,18,55])):
        c=ws.cell(3,j+1,h); style_header_dark(c); ws.column_dimensions[get_column_letter(j+1)].width=w

    regioes=[
        ["Paraná (PR)", "Londrina, Maringá, Cascavel, Pato Branco, Ponta Grossa", "100–500 km", "1–2", "Hub logístico local. Frete muito competitivo. Ideal para primeiras vendas."],
        ["Santa Catarina (SC)", "Chapecó (suinocultura/soja), Lages, Campos Novos, Canoinhas", "300–800 km", "2–3", "Forte polo de soja e milho no Oeste catarinense."],
        ["Rio Grande do Sul (RS)", "Passo Fundo, Cruz Alta, Ijuí, Santa Rosa, Pelotas (arroz)", "500–1.100 km", "2–3", "Maior produtor de arroz e expressivo em soja. Alto giro de bombas hidráulicas."],
        ["São Paulo (SP)", "Ribeirão Preto, Sertãozinho, Barretos, Araçatuba, Presidente Prudente", "400–950 km", "2–4", "Maior mercado consumidor. Cana-de-açúcar + soja. Alta demanda por sensores."],
        ["Minas Gerais (MG)", "Uberaba, Uberlândia, Patos de Minas, Montes Claros", "700–1.200 km", "3–5", "Cerrado mineiro. Café + soja. Crescimento acelerado de precisão agrícola."],
        ["Mato Grosso (MT)", "Sorriso, Lucas do Rio Verde, Rondonópolis, Sinop, Nova Mutum", "1.400–2.000 km", "4–7", "MAIOR produtor de soja do Brasil. Alto volume de máquinas e peças. Frete caro mas mercado enorme."],
        ["Mato Grosso do Sul (MS)", "Dourados, Maracaju, Sidrolândia, Campo Grande", "900–1.200 km", "3–5", "Soja, milho e pecuária. Próximo ao PR — logística mais fácil que MT."],
        ["Goiás (GO) / DF", "Rio Verde, Jataí, Cristalina, Luziânia, Brasília", "900–1.300 km", "3–5", "Rio Verde = capital nacional de grãos. Alta demanda por insumos e peças."],
        ["Bahia (BA)", "Luís Eduardo Magalhães, Barreiras, Correntina, Formosa do Rio Preto", "1.800–2.400 km", "5–8", "MATOPIBA — fronteira agrícola mais dinâmica do Brasil. Alta demanda, logística desafiadora."],
        ["Maranhão (MA) / Piauí (PI)", "Balsas, São Raimundo das Mangabeiras, Uruçuí, Bom Jesus", "2.200–2.800 km", "6–9", "MATOPIBA nordestino. Rápido crescimento — atenção especial para prazo de entrega."],
        ["Pará (PA) / Tocantins (TO)", "Santarém, Paragominas, Redenção, Gurupi, Palmas", "2.400–3.000 km", "7–10","Fronteira norte da soja. Infraestrutura limitada — transportadora especializada obrigatória."],
        ["Amazônia (AM/RO/AC/RR/AP)", "Manaus, Porto Velho, Rio Branco, Macapá", "3.000–5.000 km", "8–14","Logística complexa. Correios muitas vezes único modal. Frete pode superar 30% do valor do produto."],
    ]

    for i,reg in enumerate(regioes):
        r=i+4; ws.row_dimensions[r].height=24
        fundo=C_GRAY if i%2==0 else C_WHITE
        for j,val in enumerate(reg):
            c=ws.cell(r,j+1,val); c.border=border_thin(); c.font=font(size=9)
            c.fill=fill(C_GREEN_LIGHT if j==0 else fundo)
            c.alignment=align("left","center",wrap=True)
            if j==2: c.alignment=align("center","center")
            if j==3: c.alignment=align("center","center")


# ════════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("📊 Gerando planilhas de preços e fretes...")
    products = load_products()
    print(f"   {len(products)} produtos carregados do CSV")
    criar_planilha_precos(products)
    criar_planilha_fretes()
    print("\n🎉 Concluído! Arquivos gerados em util/")
    print("   • planilha_precos_master.xlsx  — formação de preço + exportação Shopify")
    print("   • planilha_fretes_master.xlsx  — política de fretes completa")
